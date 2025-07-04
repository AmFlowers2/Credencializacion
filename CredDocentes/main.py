import tkinter as tk, os
from tkinter import filedialog, messagebox
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from procesamiento import procesarDatosDocentes,genZip

borrador_pedido = None

class App:
    def __init__(self, master):

        print(f"En esta ventana se mostrarán los posibles errores\nY advertencias que se encuentren en los datos de los docentes\n")
        self.master = master
        master.title("Preparación de Credenciales de Docentes")
        master.geometry("450x500")
        master.configure(bg="#f0f0f0")

        titulo = tk.Label(master, text="Preparación de Credenciales Docentes", font=("Segoe UI", 16, "bold"), bg="#f0f0f0")
        titulo.grid(row=0, column=0, columnspan=3, pady=(10, 20))

        self.archivos_cargados = {
            "dfDocentesIntranet": None,
            "dfTodos": None,
            "ruta_fotos": None
        }

        self.labels_estado = {}

        self.crear_apartado("Archivo de Docentes Nuevos", "dfDocentesIntranet", 1)
        self.crear_apartado("Archivo de Todos", "dfTodos", 3)
        self.crear_apartado("Carpeta de fotos recibidas", "ruta_fotos", 5)

        estilo_boton = {"font": ("Segoe UI", 10), "bg": "#ff6961", "fg": "black", "activebackground": "#45a049"}

        #Boton para procesar los datos
        self.boton_procesar = tk.Button(master, text="Procesar", state=tk.DISABLED, command=self.procesar, **estilo_boton)
        self.boton_procesar.grid(row=7, column=0, columnspan=3, pady=(35, 20), ipadx=17, ipady=5)

        #Label para mostrar el resultado del procesamiento
        self.lbl_resultado = tk.Label(master, text="", font=("Segoe UI", 10), bg="#f0f0f0")
        self.lbl_resultado.grid(row=8, column=0, columnspan=3, pady=(5, 10))

        #Boton para generar el archivo de Excel
        self.btn_generar_excel = tk.Button(master, text="Generar Excel", state=tk.DISABLED, command=self.generar_excel, **estilo_boton)
        self.btn_generar_excel.grid(row=9, column=0, columnspan=3, pady=(5, 20), ipadx=10, ipady=5)
    
    def crear_apartado(self, texto, clave, fila):

        #Label del tipo de archivo
        label = tk.Label(self.master, text=texto + ":", font=("Segoe UI", 10), bg="#f0f0f0")
        label.grid(row=fila, column=0, sticky="w", padx=20)

        #Botón para seleccionar el archivo o carpeta
        boton = tk.Button(self.master, text="Seleccionar", command=lambda: self.seleccionar_archivo(clave), font=("Segoe UI", 9))
        boton.grid(row=fila, column=1, padx=10)

        #Label para mostrar el estado del archivo o carpeta
        estado = tk.Label(self.master, text="⛔ No cargado", font=("Segoe UI", 9), bg="#f0f0f0", fg="red")
        estado.grid(row=fila, column=2, sticky="w")

        # Label para mostrar solo el nombre del archivo
        nombre_archivo = tk.Label(self.master, text="", font=("Segoe UI", 8), bg="#f0f0f0", fg="gray")
        nombre_archivo.grid(row=fila + 1, column=0, columnspan=3, sticky="w", padx=20)

        self.labels_estado[clave] = estado
        self.labels_estado[f"{clave}_nombre"] = nombre_archivo

    def seleccionar_archivo(self, clave):
        if clave == "ruta_fotos":
            ruta = filedialog.askdirectory(title="Selecciona la carpeta de fotos")
        else:
            ruta = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        if ruta:
            self.archivos_cargados[clave] = ruta
            self.labels_estado[clave].config(text="✅ Archivo cargado", fg="green")
            nombre_archivo = os.path.basename(ruta)
            self.labels_estado[f"{clave}_nombre"].config(text=f"Has seleccionado: {nombre_archivo}")
            self.verificar_todo_cargado()

    def verificar_todo_cargado(self):
        if all(self.archivos_cargados.values()):
            estilo_boton = {"font": ("Segoe UI", 10), "bg": "#00913f", "fg": "white", "activebackground": "#114232", "state": "normal"}
            self.boton_procesar.config(**estilo_boton)

    def procesar(self):
        dfDocentesIntranet_path = self.archivos_cargados["dfDocentesIntranet"]
        dfTodos_path = self.archivos_cargados["dfTodos"]
        ruta = self.archivos_cargados["ruta_fotos"]

        global borrador_pedido
        borrador_pedido = procesarDatosDocentes(dfDocentesIntranet_path, dfTodos_path, ruta)

        if (borrador_pedido is not None) and (len(borrador_pedido) > 1):
            estilo_boton = {"font": ("Segoe UI", 10), "bg": "#00913f", "fg": "white", "activebackground": "#114232", "state": "normal"}
            self.lbl_resultado.config(text=f"Procesamiento completo: \n{len(borrador_pedido)} registros creados", fg="green")
            self.btn_generar_excel.config(**estilo_boton)
        else:
            messagebox.showerror("Error", "No se generó ningún resultado.")

    def generar_excel(self):
        #Generar el archivo de Excel para hacer el pedido de credenciales
        fecha = datetime.today().strftime('%Y %m %d') #Obtenemos la fecha del dia de hoy en formato AAAA MM DD
        nombre_excel = f"Pedido DOC {fecha}.xlsx"
        
        archivo_excel = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=nombre_excel,
            title="Guardar archivo como ..."
        )

        if archivo_excel:
            
            genZip(self.archivos_cargados["ruta_fotos"], fecha, borrador_pedido) #Genera el zip con las fotos redimensionadas

            #Convertir el borrado en un excel
            borrador_pedido.to_excel(archivo_excel,index=False, engine="openpyxl")

            #A partir de aqui son simples estilos, pero que calzan con el formato de Excel de pedidos anteriores
            #Generado con IA
            wb = load_workbook(archivo_excel)
            ws = wb.active
            # Definir estilo del encabezado
            header_font = Font(color="FFFFFF", bold=True)  # Blanco y negrita
            header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo
            # Aplicar estilo al encabezado
            for cell in ws[1]:  # Primera fila (encabezado)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="center")
            # Ajustar altura del encabezado
            ws.row_dimensions[1].height = 36
            left_align = Alignment(horizontal="left", vertical="center")
            center_align = Alignment(horizontal="center", vertical="center")
            for i,column in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = column[0].column_letter  # Obtener la letra de la columna
                align = left_align if i <= 3 else center_align
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                            cell.alignment = align
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2
            # Guardar cambios y geberar el Excel
            wb.save(archivo_excel)
            messagebox.showinfo("Éxito", "Archivo guardado correctamente")
        else:
            messagebox.showwarning("Advertencia", "No has guardado el archivo")

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()