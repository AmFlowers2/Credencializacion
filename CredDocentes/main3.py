import tkinter as tk
from tkinter import filedialog, messagebox
import os
from datetime import datetime
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from procesamiento import procesarDatosDocentes, genZip

borrador_pedido = None

def detectar_tema_sistema():
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize")
        valor, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
        if valor == 0:
            return "darkly" # Tema oscuro
        else:
            return "lumen" # Tema claro
    except Exception:
        return "lumen" # Si falla, por defecto claro

class App(ttk.Window):
    def __init__(self, tema_detectado):
        super().__init__(themename=tema_detectado)
        
        self.title("Gestor de Credenciales Docentes")
        self.geometry("750x780")

        self.archivos_cargados = {
            "dfDocentesIntranet": None,
            "dfTodos": None,
            "ruta_fotos": None
        }
        self.widgets_estado = {}

        # --- CABECERA ---
        header_frame = ttk.Frame(self, padding=20)
        header_frame.pack(fill=X)
        
        
        lbl_titulo = ttk.Label(header_frame, text="PREPARACIÓN DE CREDENCIALES DOCENTES", font=("Helvetica", 18, "bold"), bootstyle="sucess")
        lbl_titulo.pack(pady=5)

        # --- ARCHIVOS ---
        files_frame = ttk.Labelframe(self, text="  Archivos Requeridos  ", padding=15, bootstyle="info")
        files_frame.pack(fill=X, padx=20, pady=10)

        self.crear_input(files_frame, "1. Docentes Nuevos (Intranet):", "dfDocentesIntranet", 0)
        self.crear_input(files_frame, "2. Base de Datos Completa:", "dfTodos", 1)
        self.crear_input(files_frame, "3. Carpeta de Fotos:", "ruta_fotos", 2, es_carpeta=True)

        # --- ACCIONES ---
        action_frame = ttk.Frame(self, padding=20)
        action_frame.pack(fill=BOTH, expand=YES)

        # BOTÓN PROCESAR
        self.btn_procesar = ttk.Button(
            action_frame, 
            text="PROCESAR DATOS", 
            command=self.procesar, 
            state=DISABLED, 
            bootstyle="primary", 
            width=25
        )
        self.btn_procesar.pack(pady=10)

        self.lbl_resultado = ttk.Label(action_frame, text="Esperando archivos...", font=("Helvetica", 10), bootstyle="secondary")
        self.lbl_resultado.pack(pady=5)

        ttk.Separator(action_frame, orient=HORIZONTAL).pack(fill=X, pady=15)

        # BOTÓN GENERAR
        self.btn_generar = ttk.Button(
            action_frame, 
            text="GENERAR EXCEL Y ZIP", 
            command=self.generar_excel, 
            state=DISABLED, 
            bootstyle="success", 
            width=25
        )
        self.btn_generar.pack(pady=10)

    def crear_input(self, parent, texto, clave, fila, es_carpeta=False):
        frame = ttk.Frame(parent)
        frame.pack(fill=X, pady=8)

        ttk.Label(frame, text=texto, font=("Helvetica", 9, "bold")).pack(anchor=W)

        input_box = ttk.Frame(frame)
        input_box.pack(fill=X, pady=2)

        entry_var = tk.StringVar()
        entry = ttk.Entry(input_box, textvariable=entry_var, state="readonly")
        entry.pack(side=LEFT, fill=X, expand=YES, padx=(0, 10))
        self.widgets_estado[f"{clave}_entry"] = entry_var
        btn = ttk.Button(
            input_box, 
            text="Examinar", 
            command=lambda: self.seleccionar_archivo(clave, es_carpeta),
            bootstyle="primary"
        )
        btn.pack(side=RIGHT)

        lbl_status = ttk.Label(frame, text="Pendiente", font=("Arial", 9), bootstyle="danger")
        lbl_status.pack(anchor=W)
        self.widgets_estado[f"{clave}_status"] = lbl_status


    def seleccionar_archivo(self, clave, es_carpeta):
        if es_carpeta:
            ruta = filedialog.askdirectory(title="Selecciona la carpeta de fotos")
        else:
            ruta = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        
        if ruta:
            self.archivos_cargados[clave] = ruta
            nombre_archivo = os.path.basename(ruta)
            self.widgets_estado[f"{clave}_entry"].set(ruta)
            
            lbl = self.widgets_estado[f"{clave}_status"]
            lbl.config(text=f"✅ Listo: {nombre_archivo}", bootstyle="success")
            
            self.verificar_todo_cargado()


    def verificar_todo_cargado(self):
        if all(self.archivos_cargados.values()):
            self.btn_procesar.config(state=NORMAL)
            self.lbl_resultado.config(text="Archivos listos. Presiona 'Procesar'.", bootstyle="primary")


    def procesar(self):
        self.btn_procesar.config(state=DISABLED, text="Procesando...")
        self.update()

        try:
            dfIntranet = self.archivos_cargados["dfDocentesIntranet"]
            dfTodos = self.archivos_cargados["dfTodos"]
            fotos = self.archivos_cargados["ruta_fotos"]

            global borrador_pedido
            borrador_pedido = procesarDatosDocentes(dfIntranet, dfTodos, fotos)

            if (borrador_pedido is not None) and (len(borrador_pedido) >= 1):
                msg = f"¡Éxito! Se generaron {len(borrador_pedido)} registros."
                self.lbl_resultado.config(text=msg, bootstyle="success")
                self.btn_generar.config(state=NORMAL)
                messagebox.showinfo("Proceso Completo", msg)
            else:
                self.lbl_resultado.config(text="No se encontraron registros.", bootstyle="warning")
                messagebox.showwarning("Atención", "El procesamiento no generó datos nuevos.")

        except Exception as e:
            messagebox.showerror("Error Crítico", f"Ocurrió un error:\n{str(e)}")
            self.lbl_resultado.config(text="Error en el proceso", bootstyle="danger")
        
        finally:
            self.btn_procesar.config(state=NORMAL, text="PROCESAR DATOS")


    def generar_excel(self):
        fecha = datetime.today().strftime('%Y %m %d')
        nombre_defecto = f"Pedido DOC {fecha}.xlsx"
        
        archivo_excel = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=nombre_defecto,
            title="Guardar Pedido y Zip"
        )

        if not archivo_excel:
            return

        try:
            self.btn_generar.config(text="Generando...", state=DISABLED)
            self.update()

            genZip(archivo_excel, self.archivos_cargados["ruta_fotos"], fecha, borrador_pedido)
            borrador_pedido.to_excel(archivo_excel, index=False, engine="openpyxl")
            
            self.aplicar_estilos_excel(archivo_excel)

            messagebox.showinfo("Éxito Total", f"Archivos generados correctamente en:\n{os.path.dirname(archivo_excel)}")
            self.lbl_resultado.config(text="Ciclo finalizado correctamente.", bootstyle="success")

        except Exception as e:
            messagebox.showerror("Error al Guardar", f"No se pudo guardar el archivo:\n{e}")
        finally:
             self.btn_generar.config(text="GENERAR EXCEL Y ZIP", state=NORMAL)


    def aplicar_estilos_excel(self, ruta_archivo):
        wb = load_workbook(ruta_archivo)
        ws = wb.active

        header_font = Font(color="FFFFFF", bold=True)
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        
        ws.row_dimensions[1].height = 36
        
        left_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")
        
        for i, column in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = column[0].column_letter
            align = left_align if i <= 3 else center_align
            
            for cell in column:
                try:
                    if cell.value:
                        largo = len(str(cell.value))
                        if largo > max_length: max_length = largo
                        cell.alignment = align
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 2
            
        wb.save(ruta_archivo)


if __name__ == "__main__":
    tema_elegido = detectar_tema_sistema()
    app = App(tema_elegido)
    app.mainloop()