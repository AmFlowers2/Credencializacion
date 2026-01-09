import tkinter as tk
from tkinter import filedialog, messagebox
import os
import sys
from datetime import datetime
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from preparacion import procesarArchivos, genZip

borrador_pedido = None

def detectar_tema_sistema():
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize")
        valor, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
        return "darkly" if valor == 0 else "flatly"
    except Exception:
        return "flatly"

class App(ttk.Window):
    def __init__(self, tema_detectado):
        super().__init__(themename=tema_detectado)
        
        self.title("Preparación de Credenciales - Alumnos")
        self.geometry("700x850") # Aumenté un poco el alto para que quepa todo

        # Variables de estado
        self.archivos_cargados = {
            "dfAlumnosIntranet": None,
            "dfTodos": None,
            "ruta_fotos": None,
            "pdf_claves": None  # <--- Nuevo campo opcional
        }
        self.widgets_estado = {}

        # --- CABECERA ---
        header_frame = ttk.Frame(self, padding=20)
        header_frame.pack(fill=X)
        estilo_titulo = "inverse-primary" if tema_detectado == "flatly" else "primary"
        
        lbl_titulo = ttk.Label(header_frame, text=" CREDENCIALES DE ALUMNOS ", font=("Helvetica", 16, "bold"), bootstyle=estilo_titulo)
        lbl_titulo.pack(pady=5)

        # --- ARCHIVOS ---
        files_frame = ttk.Labelframe(self, text="  Archivos Requeridos y Opcionales  ", padding=15, bootstyle="info")
        files_frame.pack(fill=X, padx=20, pady=10)

        self.crear_input(files_frame, "1. Lista de Alumnos Activos (Intranet):", "dfAlumnosIntranet", 0)
        self.crear_input(files_frame, "2. Base de Datos de Todos:", "dfTodos", 1)
        self.crear_input(files_frame, "3. Carpeta de Fotos Recibidas:", "ruta_fotos", 2, tipo="carpeta")
        # El PDF es el cuarto elemento
        self.crear_input(files_frame, "4. PDF de Reposiciones (Opcional):", "pdf_claves", 3, tipo="pdf")

        # --- ACCIONES ---
        action_frame = ttk.Frame(self, padding=20)
        action_frame.pack(fill=BOTH, expand=YES)

        self.boton_procesar = ttk.Button(action_frame, text="PROCESAR DATOS", command=self.procesar, state=DISABLED, bootstyle="primary", width=25)
        self.boton_procesar.pack(pady=10)

        self.lbl_estado = ttk.Label(action_frame, text="Esperando archivos requeridos...", font=("Helvetica", 10), bootstyle="secondary")
        self.lbl_estado.pack(pady=5)

        ttk.Separator(action_frame, orient=HORIZONTAL).pack(fill=X, pady=15)

        self.btn_generar_excel = ttk.Button(action_frame, text="GENERAR EXCEL Y ZIP", command=self.generar_excel, state=DISABLED, bootstyle="success", width=25)
        self.btn_generar_excel.pack(pady=10)

    def crear_input(self, parent, texto, clave, fila, tipo="excel"):
        frame = ttk.Frame(parent)
        frame.pack(fill=X, pady=8)

        ttk.Label(frame, text=texto, font=("Helvetica", 9, "bold")).pack(anchor=W)

        input_box = ttk.Frame(frame)
        input_box.pack(fill=X, pady=2)

        entry_var = tk.StringVar()
        entry = ttk.Entry(input_box, textvariable=entry_var, state="readonly")
        entry.pack(side=LEFT, fill=X, expand=YES, padx=(0, 10))
        self.widgets_estado[f"{clave}_entry"] = entry_var

        btn = ttk.Button(input_box, text="Examinar", command=lambda: self.seleccionar_archivo(clave, tipo), bootstyle="secondary-outline")
        btn.pack(side=RIGHT)

        status_text = "⚪ Opcional" if tipo == "pdf" else "⛔ No cargado"
        status_style = "secondary" if tipo == "pdf" else "danger"
        
        lbl_status = ttk.Label(frame, text=status_text, font=("Arial", 9), bootstyle=status_style)
        lbl_status.pack(anchor=W)
        self.widgets_estado[f"{clave}_status"] = lbl_status

    def seleccionar_archivo(self, clave, tipo):
        if tipo == "carpeta":
            ruta = filedialog.askdirectory(title="Selecciona la carpeta de fotos")
        elif tipo == "pdf":
            ruta = filedialog.askopenfilename(filetypes=[("Archivos PDF", "*.pdf")])
        else:
            ruta = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        
        if ruta:
            self.archivos_cargados[clave] = ruta
            nombre_archivo = os.path.basename(ruta)
            self.widgets_estado[f"{clave}_entry"].set(ruta)
            self.widgets_estado[f"{clave}_status"].config(text=f"✅ Listo: {nombre_archivo}", bootstyle="success")
            self.verificar_todo_cargado()

    def verificar_todo_cargado(self):
        # El PDF NO es obligatorio para habilitar el botón
        requeridos = [self.archivos_cargados["dfAlumnosIntranet"], 
                      self.archivos_cargados["dfTodos"], 
                      self.archivos_cargados["ruta_fotos"]]
        
        if all(requeridos):
            self.boton_procesar.config(state=NORMAL)
            self.lbl_estado.config(text="Archivos listos. Presiona 'Procesar'.", bootstyle="primary")
    
    def procesar(self):
        self.boton_procesar.config(state=DISABLED, text="Procesando...")
        self.update()

        try:
            dfIntranet = self.archivos_cargados["dfAlumnosIntranet"]
            dfTodos = self.archivos_cargados["dfTodos"]
            rutaFotos = self.archivos_cargados["ruta_fotos"]
            rutaPdf = self.archivos_cargados["pdf_claves"]

            global borrador_pedido
            borrador_pedido = procesarArchivos(dfIntranet, dfTodos, rutaFotos, rutaPdf)

            if (borrador_pedido is not None) and (len(borrador_pedido) >= 1):
                msg = f"¡Proceso completo! {len(borrador_pedido)} registros creados."
                self.lbl_estado.config(text=msg, bootstyle="success")
                self.btn_generar_excel.config(state=NORMAL)
                messagebox.showinfo("Éxito", msg)
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error:\n{e}")
        finally:
            self.boton_procesar.config(state=NORMAL, text="PROCESAR DATOS")

    def generar_excel(self):
        fecha = datetime.today().strftime("%Y %m %d")
        nombre_excel = f"Pedido A {fecha}.xlsx"

        archivo_excel = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=nombre_excel,
            title="Guardar archivo como...",
        )

        if not archivo_excel:
            return

        try:
            self.btn_generar_excel.config(text="Generando...", state=DISABLED)
            self.update()

            genZip(archivo_excel, self.archivos_cargados["ruta_fotos"], fecha, borrador_pedido)
            borrador_pedido.to_excel(archivo_excel, index=False, engine="openpyxl")
            self.aplicar_estilos_excel(archivo_excel)
            messagebox.showinfo("Éxito", "Archivo Excel y ZIP guardados correctamente.\nNo olvides añadir las fotos de los alumnos de reposición al ZIP.")
            self.lbl_estado.config(text="Ciclo finalizado correctamente.", bootstyle="success")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar:\n{e}")
        finally:
            self.btn_generar_excel.config(text="GENERAR EXCEL Y ZIP", state=NORMAL)

    def aplicar_estilos_excel(self, ruta_archivo):
        """ Aplica el formato de encabezado rojo y autoajuste de columnas """
        wb = load_workbook(ruta_archivo)
        ws = wb.active

        header_font = Font(color="FFFFFF", bold=True)
        # Rojo corporativo para el Excel
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        
        ws.row_dimensions[1].height = 36
        
        left_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")
        
        for i, column in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = column[0].column_letter
            align = left_align if i <= 3 else center_align
            
            for cell in column:
                try:
                    if cell.value:
                        largo = len(str(cell.value))
                        if largo > max_length: max_length = largo
                        cell.alignment = align
                except: pass
            ws.column_dimensions[col_letter].width = max_length + 2
            
        wb.save(ruta_archivo)

if __name__ == "__main__":
    tema_elegido = detectar_tema_sistema()
    app = App(tema_elegido)
    app.mainloop()